import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import io
import json
import math
from datetime import datetime, timedelta
from PIL import Image, ImageDraw
import requests

import folium
from streamlit_folium import st_folium
from streamlit_js_eval import get_geolocation

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage

# ReportLab para geração de PDF ABNT & Profissional
from reportlab.lib.pagesizes import A4, portrait
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage, PageBreak, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas

# --- OCULTAR ELEMENTOS DA INTERFACE PADRÃO ---
ocultar_elementos = """
    <style>
    #MainMenu {visibility: hidden !important;}
    header {visibility: hidden !important;}
    footer {visibility: hidden !important;}
    .stAppHeader {display: none !important;}
    [data-testid="stStatusWidget"] {display: none !important;}
    button[title="Manage app"] {display: none !important;}
    div[class*="manageApp"] {display: none !important;}
    div[class*="StatusWidget"] {display: none !important;}
    iframe[src*="huggingface.co"] {display: none !important;}
    .badge-container, .hf-badge {display: none !important;}
    a[href*="huggingface.co/spaces"] {display: none !important;}
    .block-container {
        padding-top: 1.5rem !important;
        padding-bottom: 2rem !important;
    }
    </style>
"""
st.markdown(ocultar_elementos, unsafe_allow_html=True)

# --- ESTILIZAÇÃO CSS GLOBAL (TEMA LIGHT SLATE E SUAVE) ---
st.markdown("""
    <style>
    .stApp {
        background-color: #F8FAFC !important;
    }
    h1 {
        color: #0F172A !important;
        background: linear-gradient(135deg, #E0F2FE 0%, #F0F9FF 100%) !important;
        padding: 20px 24px !important;
        border-radius: 16px !important;
        border-left: 6px solid #0EA5E9 !important;
        box-shadow: 0 4px 20px rgba(14, 165, 233, 0.08) !important;
        font-weight: 800 !important;
    }
    h2, h3 {
        color: #0284C7 !important;
        font-weight: 700 !important;
    }
    div[data-testid="stVerticalBlock"] > div[data-testid="stBlock"] {
        background: #FFFFFF !important;
        padding: 20px !important;
        border-radius: 16px !important;
        box-shadow: 0 4px 15px rgba(148, 163, 184, 0.08) !important;
        border: 1px solid #E2E8F0 !important;
    }
    div[data-testid="stMetric"] {
        background: #F1F5F9 !important;
        border: 1px solid #E2E8F0 !important;
        border-radius: 14px !important;
        padding: 14px 18px !important;
        box-shadow: 0 2px 8px rgba(0, 0, 0, 0.03) !important;
        transition: all 0.2s ease-in-out;
    }
    div[data-testid="stMetric"]:hover {
        background: #E0F2FE !important;
        border-color: #38BDF8 !important;
        transform: translateY(-2px);
    }
    div[data-testid="stMetricLabel"] > label {
        color: #64748B !important;
        font-size: 0.85rem !important;
        font-weight: 600 !important;
    }
    div[data-testid="stMetricValue"] > div {
        color: #0369A1 !important;
        font-weight: 800 !important;
    }
    .stButton > button {
        background: linear-gradient(135deg, #0284C7 0%, #0369A1 100%) !important;
        color: white !important;
        border: none !important;
        border-radius: 10px !important;
        padding: 10px 24px !important;
        font-weight: 700 !important;
        width: 100%;
    }
    .stTextInput input, .stSelectbox select, .stNumberInput input {
        background-color: #F8FAFC !important;
        border: 1px solid #CBD5E1 !important;
        color: #1E293B !important;
        border-radius: 8px !important;
    }
    </style>
""", unsafe_allow_html=True)

# --- GERENCIAMENTO DE SESSÃO / LOGIN ---
if "logado" not in st.session_state:
    st.session_state["logado"] = False

USUARIOS = {
    "admin": "1234",
    "natanael": "sondagem2026"
}

def tela_login():
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown("## 🔐 Acesso ao Sistema")
        st.markdown("Entre com suas credenciais para acessar o Boletim de Sondagem.")
        
        usuario = st.text_input("Usuário", placeholder="Digite seu usuário")
        senha = st.text_input("Senha", type="password", placeholder="Digite sua senha")
        
        if st.button("Entrar"):
            if usuario in USUARIOS and USUARIOS[usuario] == senha:
                st.session_state["logado"] = True
                st.session_state["usuario_atual"] = usuario
                st.success("Login efetuado com sucesso!")
                st.rerun()
            else:
                st.error("Usuário ou senha incorretos.")

if not st.session_state["logado"]:
    tela_login()
    st.stop()

# --- BARRA LATERAL ---
with st.sidebar:
    st.markdown(f"👤 **Usuário:** `{st.session_state.get('usuario_atual', 'Usuário')}`")
    if st.button("🔒 Sair / Logout"):
        st.session_state["logado"] = False
        st.rerun()

# --- FUNÇÕES AUXILIARES ---
DADOS_LITOLOGIA = {
    'Solo / Cobertura':        {'cor': '#E5D3B3', 'hatch': '....'},
    'Siltito / Argilito':      {'cor': '#D2B48C', 'hatch': '----'},
    'Quartzito':               {'cor': '#FFF8DC', 'hatch': '////'},
    'Schisto / Filito':        {'cor': '#94A3B8', 'hatch': '\\\\\\\\'},
    'Gnaisse / Granito':       {'cor': '#E2E8F0', 'hatch': '++++'},
    'Basalto / Diabásio':      {'cor': '#475569', 'hatch': 'xxxx'},
    'Minério de Ferro / BIF':  {'cor': '#991B1B', 'hatch': '||||'},
    'Calcário / Dolomito':     {'cor': '#BAE6FD', 'hatch': 'OOOO'},
    'Outro':                   {'cor': '#CBD5E1', 'hatch': ''}
}

if 'manobras' not in st.session_state:
    st.session_state['manobras'] = []

# --- APLICAÇÃO PRINCIPAL ---
st.title("📋 Boletim Digital de Sondagem Mineral")
st.markdown("---")

st.header("1. Cabeçalho do Projeto & Equipe Técnica")

col_logo, col_gest = st.columns([1, 3])
with col_logo:
    st.subheader("🖼️ Logomarca da Empresa")
    logo_file = st.file_uploader("Carregar Logo (PNG/JPG)", type=['png', 'jpg', 'jpeg'])
    img_logo_pil = Image.open(logo_file) if logo_file else None
    if img_logo_pil:
        st.image(img_logo_pil, width=180)

with col_gest:
    col_g1, col_g2, col_g3 = st.columns(3)
    with col_g1:
        empresa = st.text_input("Empresa / Mineradora", value="Mineração Picuí S.A.")
        projeto = st.text_input("Nome do Projeto", value="Projeto Picuí")
    with col_g2:
        coordenador = st.text_input("Coordenador do Projeto", value="Eng. Carlos Andrade")
        supervisor = st.text_input("Supervisor de Campo", value="Téc. Roberto Lima")
    with col_g3:
        geologo = st.text_input("Geólogo Responsável", value="Geól. Mariana Costa")
        sondador = st.text_input("Sondador / Equipe", value="Natanael & Equipe")

col_furo1, col_furo2 = st.columns(2)
with col_furo1:
    furo_id = st.text_input("ID do Furo", value="F-001")
with col_furo2:
    diametro = st.selectbox("Diâmetro", ['HQ (63.5mm)', 'NQ (47.6mm)', 'BQ (36.5mm)', 'RC (Circ. Reversa)', 'Outro'])

# --- PARÂMETROS GEOGRÁFICOS E DATAS (Escopo Principal) ---
col_geo1, col_geo2, col_geo3, col_geo4 = st.columns(4)

with col_geo1:
    lat_padrao = -6.515831
    lon_padrao = -36.344525
    if 'lat_gps' not in st.session_state:
        st.session_state['lat_gps'] = lat_padrao
    if 'lon_gps' not in st.session_state:
        st.session_state['lon_gps'] = lon_padrao

    lat_furo = st.number_input("Latitude", value=st.session_state['lat_gps'], format="%.6f", key="input_lat")
    lon_furo = st.number_input("Longitude", value=st.session_state['lon_gps'], format="%.6f", key="input_lon")
    st.session_state['lat_gps'] = lat_furo
    st.session_state['lon_gps'] = lon_furo

with col_geo2:
    datum = st.text_input("Datum", value="SIRGAS 2000")
    cota_z = st.number_input("Elevação Z (m)", value=0.0, step=0.5)

with col_geo3:
    inclinacao = st.number_input("Inclinação (°)", value=-90.0, format="%.1f")
    azimute = st.number_input("Azimute (°)", value=0.0, format="%.1f")

with col_geo4:
    # Definidas no escopo principal para ficarem visíveis no Excel/PDF
    data_inicio = st.date_input("Data de Início", value=datetime.now())
    data_fim = st.date_input("Data de Término", value=datetime.now())

# --- EXPANDER DO MAPA E GPS ---
with st.expander("🌐 Capturar GPS Automático e Ver Mapa", expanded=False):
    st.markdown("#### 🎯 Captura Automática de Localização")
    
    loc = get_geolocation()
    if loc and 'coords' in loc:
        nova_lat = round(loc['coords']['latitude'], 6)
        nova_lon = round(loc['coords']['longitude'], 6)
        
        if st.session_state['lat_gps'] != nova_lat or st.session_state['lon_gps'] != nova_lon:
            st.session_state['lat_gps'] = nova_lat
            st.session_state['lon_gps'] = nova_lon
            st.success("✅ Coordenadas capturadas com sucesso!")
            st.rerun()

    m = folium.Map(location=[st.session_state['lat_gps'], st.session_state['lon_gps']], zoom_start=18, tiles=None)
    folium.TileLayer(
        tiles='https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}',
        attr='Esri World Imagery', name='Satélite (Esri)', overlay=False
    ).add_to(m)

    folium.Marker(
        [st.session_state['lat_gps'], st.session_state['lon_gps']], 
        popup=f"Furo: {furo_id}", 
        tooltip="Sua Localização Atual",
        icon=folium.Icon(color='red', icon='info-sign')
    ).add_to(m)

    st_folium(m, width="100%", height=350, key="mapa_gps")

# --- SEÇÃO 2: REGISTRO DE MANOBRAS ---
st.header("2. Registro de Manobras e Fotos do Testemunho")

if st.session_state['manobras']:
    prox_de = st.session_state['manobras'][-1]['Para (m)']
    rec_total_ant = st.session_state['manobras'][-1]['Rec. Total (m)']
else:
    prox_de = 0.0
    rec_total_ant = 0.0

prox_para = round(prox_de + 1.5, 2)

st.subheader("🛠️ Peça de Corte e Revestimento")
col_pc1, col_pc2, col_pc3, col_pc4, col_pc5 = st.columns(5)
with col_pc1:
    peca_diam = st.text_input("Diâm. Peça", value="NQ")
with col_pc2:
    peca_coroa = st.text_input("Coroa nº", placeholder="Ex: 89173-17")
with col_pc3:
    peca_calib = st.text_input("Calib. nº", placeholder="Ex: 1381/17")
with col_pc4:
    num_caixa = st.number_input("Nº da Caixa", min_value=1, value=1, step=1)
with col_pc5:
    revest_info = st.text_input("Revestimento", placeholder="Ex: HQ De 0,00 até 34,40m")

st.markdown("---")

col_m1, col_m2, col_m3, col_m4, col_m5 = st.columns(5)
with col_m1:
    de = st.number_input("De (m)", value=float(prox_de), step=0.5, format="%.2f")
with col_m2:
    para = st.number_input("Para (m)", value=float(prox_para), step=0.5, format="%.2f")
with col_m3:
    rec = st.number_input("Rec. (m)", value=round(para - de, 2), step=0.1, format="%.2f")
with col_m4:
    rec_total = st.number_input("Rec. Total Acum. (m)", value=round(rec_total_ant + rec, 2), step=0.1, format="%.2f")
with col_m5:
    rqd = st.number_input("RQD (m)", value=round((para - de) * 0.8, 2), step=0.1, format="%.2f")

col_h1, col_h2, col_h3, col_h4 = st.columns(4)
with col_h1:
    hora_ini = st.time_input("Horário Inicial", value=datetime.now().time())
with col_h2:
    hora_fim = st.time_input("Horário Final", value=datetime.now().time())
with col_h3:
    tempo_refeicao = st.text_input("Refeição", placeholder="Ex: 01:00")
with col_h4:
    manutencao_prev = st.text_input("Manutenção Preventiva", placeholder="Ex: 00:15")

col_l1, col_l2 = st.columns(2)
with col_l1:
    litologia = st.selectbox("Litologia", list(DADOS_LITOLOGIA.keys()))
with col_l2:
    alteracao = st.selectbox("Alteração", ['Solo / Inconsol.', 'Completamente Alterada', 'Muito Alterada', 'Moderadamente Alterada', 'Pouco Alterada', 'Rocha Sã'])

st.subheader("📷 Registro Fotográfico da Amostra / Caixa")
aba_cam, aba_up = st.tabs(["📸 Tirar Foto Agora", "📁 Carregar da Galeria"])

img_capturada = None
with aba_cam:
    foto_cam = st.camera_input("Tirar foto da caixa de testemunho")
    if foto_cam: img_capturada = Image.open(foto_cam)

with aba_up:
    foto_file = st.file_uploader("Selecione uma imagem", type=['jpg', 'jpeg', 'png'])
    if foto_file and not img_capturada: img_capturada = Image.open(foto_file)

col_btn1, col_btn2 = st.columns([1, 4])
with col_btn1:
    btn_adicionar = st.button("➕ Adicionar Manobra", type="primary")
with col_btn2:
    btn_remover = st.button("🗑️ Remover Última")

if btn_adicionar:
    avanco = round(para - de, 2)
    if avanco <= 0:
        st.error("⚠️ O valor 'Para' deve ser maior que 'De'!")
    else:
        pct_rec = min(100.0, round((rec / avanco) * 100, 1)) if avanco > 0 else 0.0
        pct_rqd = min(100.0, round((rqd / avanco) * 100, 1)) if avanco > 0 else 0.0
        
        if pct_rqd < 25: rqd_class = 'Muito Pobre'
        elif pct_rqd < 50: rqd_class = 'Pobre'
        elif pct_rqd < 75: rqd_class = 'Razoável'
        elif pct_rqd < 90: rqd_class = 'Boa'
        else: rqd_class = 'Excelente'

        t_ini = datetime.combine(datetime.today(), hora_ini)
        t_fim = datetime.combine(datetime.today(), hora_fim)
        if t_fim < t_ini:
            t_fim += timedelta(days=1)
        duracao_horas = round((t_fim - t_ini).total_seconds() / 3600.0, 2)
        if duracao_horas == 0:
            duracao_horas = 0.5

        st.session_state['manobras'].append({
            'Manobra': len(st.session_state['manobras']) + 1,
            'De (m)': de, 'Para (m)': para, 'Avanço (m)': avanco,
            'Rec. (m)': rec, 'Rec. Total (m)': rec_total, 'Rec (%)': pct_rec, 
            'RQD (m)': rqd, 'RQD (%)': pct_rqd, 'Qualidade RQD': rqd_class,
            'Diâm. Peça': peca_diam, 'Coroa nº': peca_coroa, 'Calib. nº': peca_calib,
            'Nº Caixa': num_caixa, 'Revestimento': revest_info,
            'Hora Inicial': hora_ini.strftime("%H:%M"), 
            'Hora Final': hora_fim.strftime("%H:%M"),
            'Duração (h)': duracao_horas,
            'Refeição': tempo_refeicao,
            'Manutenção Preventiva': manutencao_prev,
            'Litologia': litologia, 'Alteração': alteracao, 
            'Foto': img_capturada
        })
        st.success("✅ Manobra registrada!")
        st.rerun()

if btn_remover and st.session_state['manobras']:
    st.session_state['manobras'].pop()
    st.warning("🗑️ Última manobra removida.")
    st.rerun()

st.markdown("---")

# --- SEÇÃO 3: PERFIL E VISUALIZAÇÃO ---
st.header("3. Perfil Litológico, Observações Gerais e Relatórios")

obs_gerais_furo = st.text_area(
    "📝 Observações Técnicas Gerais / Notas de Campo do Furo", 
    placeholder="Digite observações importantes sobre o furo, trocas de ferramenta, perdas de água, etc...",
    height=100
)

if st.session_state['manobras']:
    df_manobras = pd.DataFrame(st.session_state['manobras'])

    plt.rcParams['hatch.linewidth'] = 1.2
    plt.rcParams['hatch.color'] = '#333333'
    fig, (ax_lito, ax_rqd, ax_rec) = plt.subplots(1, 3, figsize=(11, 4.5), sharey=True, gridspec_kw={'width_ratios': [1.3, 2, 2]})
    
    prof_max = df_manobras['Para (m)'].max()
    ax_lito.set_ylim(prof_max, 0)
    litos_usadas = set()

    for _, row in df_manobras.iterrows():
        de_m, para_m, lito = row['De (m)'], row['Para (m)'], row['Litologia']
        rqd_val, rec_val = row['RQD (%)'], row['Rec (%)']
        info_lito = DADOS_LITOLOGIA.get(lito, {'cor': '#808080', 'hatch': ''})
        litos_usadas.add(lito)

        rect = mpatches.Rectangle((0, de_m), 1, para_m - de_m, facecolor=info_lito['cor'], hatch=info_lito['hatch'], edgecolor='#1E293B', linewidth=1.2)
        ax_lito.add_patch(rect)
        ax_lito.axhline(para_m, color='#0F172A', linestyle='--', linewidth=0.8)
        ax_lito.text(0.5, (de_m + para_m)/2, f"{lito}\n({de_m:.1f}m - {para_m:.1f}m)", ha='center', va='center', fontsize=8, fontweight='bold', bbox=dict(boxstyle='round,pad=0.2', facecolor='#FFFFFF', alpha=0.85, edgecolor='#94A3B8'))

        color_rqd = '#EF4444' if rqd_val < 25 else '#F97316' if rqd_val < 50 else '#EAB308' if rqd_val < 75 else '#3B82F6' if rqd_val < 90 else '#22C55E'
        ax_rqd.barh(y=de_m + (para_m - de_m)/2, width=rqd_val, height=(para_m - de_m)*0.8, color=color_rqd, edgecolor='black', linewidth=0.8)
        ax_rec.barh(y=de_m + (para_m - de_m)/2, width=rec_val, height=(para_m - de_m)*0.8, color='#0284C7', edgecolor='black', linewidth=0.8)

    ax_lito.set_xlim(0, 1)
    ax_lito.set_title("Estratigrafia", fontsize=10, fontweight='bold')
    ax_lito.set_ylabel("Profundidade (m)", fontsize=9, fontweight='bold')
    ax_lito.get_xaxis().set_visible(False)
    ax_rqd.set_xlim(0, 105)
    ax_rqd.set_title("RQD (%)", fontsize=10, fontweight='bold')
    ax_rec.set_xlim(0, 105)
    ax_rec.set_title("Recuperação (%)", fontsize=10, fontweight='bold')
    plt.tight_layout()

    st.pyplot(fig)
    
    img_perfil_bytes = io.BytesIO()
    fig.savefig(img_perfil_bytes, format='png', dpi=300, bbox_inches='tight')
    plt.close(fig)

    st.dataframe(df_manobras.drop(columns=['Foto']), use_container_width=True, hide_index=True)

    # --- SEÇÃO 4: DASHBOARD DE PRODUÇÃO ---
    st.markdown("---")
    st.header("⚡ 4. Dashboard de Produção & Indicadores")

    avanco_total = df_manobras['Avanço (m)'].sum()
    df_manobras['Duração (h)'] = [m.get('Duração (h)', 0.5) for m in st.session_state['manobras']]
    tempo_total_h = df_manobras['Duração (h)'].sum()

    taxa_perf_media = round(avanco_total / tempo_total_h, 2) if tempo_total_h > 0 else 0.0
    rec_media = round(df_manobras['Rec (%)'].mean(), 1)
    rqd_medio = round(df_manobras['RQD (%)'].mean(), 1)

    kpi1, kpi2, kpi3, kpi4, kpi5 = st.columns(5)
    kpi1.metric("Avanço Acumulado", f"{avanco_total:.2f} m")
    kpi2.metric("Horas de Operação", f"{tempo_total_h:.2f} h")
    kpi3.metric("Rendimento Médio", f"{taxa_perf_media:.2f} m/h")
    kpi4.metric("Recuperação Média", f"{rec_media:.1f} %")
    kpi5.metric("RQD Médio", f"{rqd_medio:.1f} %")

    st.markdown("<br>", unsafe_allow_html=True)

    col_dash1, col_dash2 = st.columns(2)
    with col_dash1:
        st.subheader("📈 Progresso do Avanço")
        fig_dash1, ax_d1 = plt.subplots(figsize=(5, 3.5), facecolor='#FFFFFF')
        ax_d1.set_facecolor('#F8FAFC')
        
        ax_d1.plot(df_manobras['Manobra'], df_manobras['Avanço (m)'], marker='o', color='#0284C7', linewidth=2, label='Avanço (m)')
        ax_d1.plot(df_manobras['Manobra'], df_manobras['Para (m)'], marker='s', color='#E11D48', linestyle='--', linewidth=2, label='Profundidade (m)')
        
        ax_d1.set_xlabel("Manobra", color='#475569')
        ax_d1.set_ylabel("Metros", color='#475569')
        ax_d1.grid(True, color='#E2E8F0', linestyle=':', alpha=0.8)
        ax_d1.legend(facecolor='#FFFFFF', edgecolor='#CBD5E1')
        plt.tight_layout()
        st.pyplot(fig_dash1)
        plt.close(fig_dash1)

    with col_dash2:
        st.subheader("🌋 Distribuição Litológica")
        lito_dist = df_manobras.groupby('Litologia')['Avanço (m)'].sum()
        fig_dash2, ax_d2 = plt.subplots(figsize=(5, 3.5), facecolor='#FFFFFF')
        ax_d2.set_facecolor('#F8FAFC')
        
        cores_suaves = ['#0EA5E9', '#F43F5E', '#10B981', '#F59E0B', '#8B5CF6', '#EC4899']
        ax_d2.pie(
            lito_dist, 
            labels=lito_dist.index, 
            autopct='%1.1f%%', 
            colors=cores_suaves[:len(lito_dist)],
            startangle=140, 
            wedgeprops=dict(width=0.4, edgecolor='#FFFFFF', linewidth=2)
        )
        ax_d2.set_title("Participação na Metragem", color='#475569', fontsize=9)
        plt.tight_layout()
        st.pyplot(fig_dash2)
        plt.close(fig_dash2)

    st.markdown("---")
    st.markdown("### ✍️ Validação Técnica")
    st.info(f"O documento gerado conterá um campo para assinatura física/manual do **{geologo}**.")

    col_exp1, col_exp2 = st.columns(2)

    # --- EXPORTAÇÃO EXCEL ---
    with col_exp1:
        buffer_xls = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Boletim de Sondagem"
        ws.views.sheetView[0].showGridLines = True

        font_titulo = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
        font_sub = Font(name='Calibri', size=10, italic=True, color='FFFFFF')
        font_sec = Font(name='Calibri', size=11, bold=True, color='0F172A')
        font_header = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        font_body = Font(name='Calibri', size=10)
        font_total = Font(name='Calibri', size=10, bold=True)

        fill_banner = PatternFill(start_color='0284C7', end_color='0284C7', fill_type='solid')
        fill_sec = PatternFill(start_color='E0F2FE', end_color='E0F2FE', fill_type='solid')
        fill_header = PatternFill(start_color='0369A1', end_color='0369A1', fill_type='solid')
        fill_zebra = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
        fill_total = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')

        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )
        double_bottom = Border(
            top=Side(style='thin', color='0F172A'),
            bottom=Side(style='double', color='0F172A')
        )

        align_center = Alignment(horizontal='center', vertical='center')
        align_left = Alignment(horizontal='left', vertical='center')
        align_right = Alignment(horizontal='right', vertical='center')

        if img_logo_pil:
            img_logo_excel_buf = io.BytesIO()
            img_logo_pil.save(img_logo_excel_buf, format='PNG')
            img_logo_excel_buf.seek(0)
            xl_logo = OpenpyxlImage(img_logo_excel_buf)
            xl_logo.width = 110
            xl_logo.height = 40
            ws.add_image(xl_logo, 'A1')

        ws.merge_cells('C1:L1')
        ws['C1'] = empresa.upper()
        ws['C1'].font = font_titulo
        ws['C1'].fill = fill_banner
        ws['C1'].alignment = align_center

        ws.merge_cells('A2:L2')
        ws['A2'] = f"BOLETIM TÉCNICO DE SONDAGEM GEOLÓGICA - FURO {furo_id}"
        ws['A2'].font = font_sub
        ws['A2'].fill = fill_banner
        ws['A2'].alignment = align_center

        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 18

        ws.merge_cells('A4:L4')
        ws['A4'] = "1. DADOS DE GESTÃO E LOCALIZAÇÃO"
        ws['A4'].font = font_sec
        ws['A4'].fill = fill_sec
        ws['A4'].alignment = align_left

        dados_header = [
            [("Projeto:", projeto), ("Coordenador:", coordenador), ("Latitude:", lat_furo)],
            [("ID Furo:", furo_id), ("Supervisor:", supervisor), ("Longitude:", lon_furo)],
            [("Diâmetro:", diametro), ("Geólogo Resp.:", geologo), ("Início:", str(data_inicio))],
            [("Inclin./Az.:", f"{inclinacao}° / {azimute}°"), ("Sondador:", sondador), ("Datum:", datum)]
        ]

        curr_row = 5
        for row in dados_header:
            col_pairs = [(1,2,3), (4,5,6), (7,8,9)]
            for idx, (lbl, val) in enumerate(row):
                c_lbl, c_val_start, c_val_end = col_pairs[idx]
                ws.cell(row=curr_row, column=c_lbl, value=lbl).font = Font(name='Calibri', size=10, bold=True)
                ws.cell(row=curr_row, column=c_lbl).alignment = align_left
                
                if c_val_start != c_val_end:
                    ws.merge_cells(start_row=curr_row, start_column=c_val_start, end_row=curr_row, end_column=c_val_end)
                cell_v = ws.cell(row=curr_row, column=c_val_start, value=val)
                cell_v.font = font_body
                cell_v.alignment = align_left
            curr_row += 1

        curr_row += 1
        ws.merge_cells(f'A{curr_row}:L{curr_row}')
        ws[f'A{curr_row}'] = "2. REGISTRO DE MANOBRAS E PARÂMETROS GEOTÉCNICOS"
        ws[f'A{curr_row}'].font = font_sec
        ws[f'A{curr_row}'].fill = fill_sec
        ws[f'A{curr_row}'].alignment = align_left

        curr_row += 1
        df_excel = df_manobras.drop(columns=['Foto'])
        
        for c_idx, col_name in enumerate(df_excel.columns, 1):
            cell = ws.cell(row=curr_row, column=c_idx, value=col_name)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = thin_border
        ws.row_dimensions[curr_row].height = 22

        header_row_idx = curr_row
        curr_row += 1
        for r_idx, row in df_excel.iterrows():
            row_fill = fill_zebra if r_idx % 2 == 1 else PatternFill(fill_type=None)
            for c_idx, val in enumerate(row.values, 1):
                cell = ws.cell(row=curr_row, column=c_idx, value=val)
                cell.font = font_body
                cell.border = thin_border
                cell.fill = row_fill
                
                if isinstance(val, (int, float)):
                    cell.alignment = align_right
                    if "Rec (%)" in df_excel.columns[c_idx-1] or "RQD (%)" in df_excel.columns[c_idx-1]:
                        cell.number_format = '0.0'
                    elif "m" in df_excel.columns[c_idx-1]:
                        cell.number_format = '0.00'
                else:
                    cell.alignment = align_center if c_idx == 1 else align_left
            curr_row += 1

        ws.cell(row=curr_row, column=1, value="Total / Média").font = font_total
        ws.cell(row=curr_row, column=1).alignment = align_center
        ws.cell(row=curr_row, column=1).fill = fill_total
        ws.cell(row=curr_row, column=1).border = double_bottom

        for c_idx in range(2, len(df_excel.columns) + 1):
            col_name = df_excel.columns[c_idx-1]
            cell = ws.cell(row=curr_row, column=c_idx)
            cell.font = font_total
            cell.fill = fill_total
            cell.border = double_bottom
            cell.alignment = align_right

            start_letter = get_column_letter(c_idx)
            start_cell = f"{start_letter}{header_row_idx + 1}"
            end_cell = f"{start_letter}{curr_row - 1}"

            if col_name in ['Avanço (m)', 'Rec. (m)', 'RQD (m)']:
                cell.value = f"=SUM({start_cell}:{end_cell})"
                cell.number_format = '0.00'
            elif col_name in ['Rec (%)', 'RQD (%)']:
                cell.value = f"=AVERAGE({start_cell}:{end_cell})"
                cell.number_format = '0.0'

        wb.save(buffer_xls)
        buffer_xls.seek(0)
        
        st.download_button(
            label="📊 Baixar Planilha Excel Completa",
            data=buffer_xls,
            file_name=f"boletim_sondagem_{furo_id}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

    with col_exp2:
        st.download_button(
            label="📄 Exportar Relatório em PDF",
            data=buffer_xls,
            file_name=f"boletim_sondagem_{furo_id}.pdf",
            mime="application/pdf",
            use_container_width=True
        )
