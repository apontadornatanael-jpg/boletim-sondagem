import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import io
import json
import math
from datetime import datetime
from PIL import Image, ImageDraw
import requests

import folium
from streamlit_folium import st_folium

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage

# ReportLab para geração de PDF ABNT & Profissional
from reportlab.lib.pagesizes import A4, portrait
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage, PageBreak, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas

# --- OCULTAR ELEMENTOS PADRÃO DO STREAMLIT ---
ocultar_elementos = """
    <style>
    #MainMenu {visibility: hidden !important;}
    header {visibility: hidden !important;}
    footer {visibility: hidden !important;}
    .stAppHeader {display: none !important;}
    [data-testid="stStatusWidget"] {display: none !important;}
    button[title="Manage app"] {display: none !important;}
    div[class*="manageApp"] {display: none !important;}
    div[class*="StatusWidget"] {display: none !important;}
    iframe[src*="huggingface.co"] {display: none !important;}
    .badge-container, .hf-badge {display: none !important;}
    a[href*="huggingface.co/spaces"] {display: none !important;}
    .block-container {
        padding-top: 1rem !important;
        padding-bottom: 1rem !important;
    }
    </style>
"""
st.markdown(ocultar_elementos, unsafe_allow_html=True)

# --- ESTILIZAÇÃO CUSTOMIZADA ---
st.markdown("""
    <style>
    div[data-testid="stVerticalBlock"] > div[data-testid="stBlock"] {
        background: #FFFFFF;
        padding: 20px;
        border-radius: 16px;
        box-shadow: 0 4px 15px rgba(0, 0, 0, 0.05);
        border: 1px solid #E2E8F0;
    }
    h1 {
        color: #0F172A !important;
        background: linear-gradient(135deg, #E0F2FE 0%, #F0F9FF 100%);
        padding: 16px 20px;
        border-radius: 14px;
        border-left: 6px solid #0284C7;
        font-weight: 800 !important;
    }
    h2, h3 {
        color: #0369A1 !important;
        font-weight: 700 !important;
    }
    .stButton > button {
        background: linear-gradient(135deg, #0284C7 0%, #0369A1 100%) !important;
        color: white !important;
        border: none !important;
        border-radius: 10px !important;
        padding: 10px 24px !important;
        font-weight: 700 !important;
        width: 100%;
    }
    </style>
""", unsafe_allow_html=True)

# --- INICIALIZAÇÃO DE SESSÃO & BANCO EM MEMÓRIA ---
if "logado" not in st.session_state:
    st.session_state["logado"] = False

if 'manobras' not in st.session_state:
    st.session_state['manobras'] = []

# Inicialização de variáveis globais na sessão para compartilhar entre abas/páginas
if 'dados_projeto' not in st.session_state:
    st.session_state['dados_projeto'] = {
        'empresa': 'Mineração Picuí S.A.',
        'projeto': 'Projeto Picuí',
        'coordenador': 'Eng. Carlos Andrade',
        'supervisor': 'Téc. Roberto Lima',
        'geologo': 'Geól. Mariana Costa',
        'sondador': 'Natanael & Equipe',
        'furo_id': 'F-001',
        'diametro': 'HQ (63.5mm)',
        'lat_gps': -6.515831,
        'lon_gps': -36.344525,
        'datum': 'SIRGAS 2000',
        'inclinacao': -90.0,
        'azimute': 0.0,
        'data_inicio': datetime.now().date(),
        'data_fim': datetime.now().date(),
        'obs_gerais': '',
        'logo_img': None
    }

USUARIOS = {
    "admin": "1234",
    "natanael": "sondagem2026"
}

DADOS_LITOLOGIA = {
    'Solo / Cobertura':        {'cor': '#E5D3B3', 'hatch': '....'},
    'Siltito / Argilito':      {'cor': '#D2B48C', 'hatch': '----'},
    'Quartzito':                {'cor': '#FFF8DC', 'hatch': '////'},
    'Schisto / Filito':        {'cor': '#94A3B8', 'hatch': '\\\\\\\\'},
    'Gnaisse / Granito':       {'cor': '#E2E8F0', 'hatch': '++++'},
    'Basalto / Diabásio':      {'cor': '#475569', 'hatch': 'xxxx'},
    'Minério de Ferro / BIF': {'cor': '#991B1B', 'hatch': '||||'},
    'Calcário / Dolomito':     {'cor': '#BAE6FD', 'hatch': 'OOOO'},
    'Outro':                   {'cor': '#CBD5E1', 'hatch': ''}
}

# --- FUNÇÃO DE OBTENÇÃO DE MAPA SATÉLITE ESRI PARA O PDF ---
def obter_mapa_satelite_esri_alta_res(lat, lon, zoom=16, width=1200, height=600):
    try:
        n = 2.0 ** zoom
        lat_rad = math.radians(lat)
        x_exact = (lon + 180.0) / 360.0 * n
        y_exact = (1.0 - math.log(math.tan(lat_rad) + (1.0 / math.cos(lat_rad))) / math.pi) / 2.0 * n
        
        x_center_tile = int(math.floor(x_exact))
        y_center_tile = int(math.floor(y_exact))
        
        x_offset = int((x_exact - x_center_tile) * 256)
        y_offset = int((y_exact - y_center_tile) * 256)
        
        canvas_img = Image.new('RGBA', (768, 768))
        headers = {'User-Agent': 'Mozilla/5.0'}
        
        for dx in range(-1, 2):
            for dy in range(-1, 2):
                tx = x_center_tile + dx
                ty = y_center_tile + dy
                url = f"https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{zoom}/{ty}/{tx}"
                res = requests.get(url, headers=headers, timeout=5)
                if res.status_code == 200:
                    tile_img = Image.open(io.BytesIO(res.content)).convert("RGBA")
                    canvas_img.paste(tile_img, ((dx + 1) * 256, (dy + 1) * 256))
        
        px_center_x = 256 + x_offset
        px_center_y = 256 + y_offset
        
        crop_w, crop_h = 600, 300
        left = max(0, px_center_x - crop_w // 2)
        top = max(0, px_center_y - crop_h // 2)
        right = left + crop_w
        bottom = top + crop_h
        
        cropped_img = canvas_img.crop((left, top, right, bottom))
        cropped_img = cropped_img.resize((width, height), Image.Resampling.LANCZOS)
        
        draw = ImageDraw.Draw(cropped_img)
        cx, cy = width // 2, height // 2
        
        r = 14
        draw.ellipse((cx - r, cy - r*2, cx + r, cy), fill="#E11D48", outline="#FFFFFF", width=3)
        draw.polygon([(cx - r + 2, cy - r//2), (cx + r - 2, cy - r//2), (cx, cy + r//2)], fill="#E11D48")
        draw.ellipse((cx - r//2, cy - r*1.5, cx + r//2, cy - r*0.5), fill="#FFFFFF")
        
        img_out = io.BytesIO()
        cropped_img.save(img_out, format="PNG", dpi=(300, 300))
        img_out.seek(0)
        return img_out
    except Exception:
        return None

# --- TELA DE LOGIN ---
def tela_login():
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown("## 🔐 Acesso ao Sistema")
        st.markdown("Entre com suas credenciais para acessar o Boletim de Sondagem.")
        
        usuario = st.text_input("Usuário", placeholder="Digite seu usuário")
        senha = st.text_input("Senha", type="password", placeholder="Digite sua senha")
        
        if st.button("Entrar"):
            if usuario in USUARIOS and USUARIOS[usuario] == senha:
                st.session_state["logado"] = True
                st.session_state["usuario_atual"] = usuario
                st.success("Login efetuado com sucesso!")
                st.rerun()
            else:
                st.error("Usuário ou senha incorretos.")

if not st.session_state["logado"]:
    tela_login()
    st.stop()

# --- NAVEGAÇÃO LATERAL (SHELL STYLE) ---
with st.sidebar:
    st.title("⛏️ Sondagem Digital")
    st.markdown(f"👤 **Usuário:** `{st.session_state.get('usuario_atual', 'Usuário')}`")
    st.markdown("---")
    
    menu_opcao = st.radio(
        "Navegação do Sistema:",
        [
            "📌 1. Projeto e Coordenadas",
            "🛠️ 2. Cadastro de Manobras",
            "📊 3. Perfil Geológico",
            "📄 4. Exportação & Relatório PDF"
        ]
    )
    
    st.markdown("---")
    if st.button("🔒 Sair / Logout"):
        st.session_state["logado"] = False
        st.rerun()

# --- PÁGINA 1: PROJETO E COORDENADAS ---
if menu_opcao == "📌 1. Projeto e Coordenadas":
    st.title("📋 Configurações do Projeto e Localização")
    
    col_logo, col_gest = st.columns([1, 3])
    with col_logo:
        st.subheader("🖼️ Logomarca")
        logo_file = st.file_uploader("Carregar Logo (PNG/JPG)", type=['png', 'jpg', 'jpeg'])
        if logo_file:
            st.session_state['dados_projeto']['logo_img'] = Image.open(logo_file)
        if st.session_state['dados_projeto']['logo_img']:
            st.image(st.session_state['dados_projeto']['logo_img'], width=180)

    with col_gest:
        col_g1, col_g2, col_g3 = st.columns(3)
        with col_g1:
            st.session_state['dados_projeto']['empresa'] = st.text_input("Empresa / Mineradora", value=st.session_state['dados_projeto']['empresa'])
            st.session_state['dados_projeto']['projeto'] = st.text_input("Nome do Projeto", value=st.session_state['dados_projeto']['projeto'])
        with col_g2:
            st.session_state['dados_projeto']['coordenador'] = st.text_input("Coordenador do Projeto", value=st.session_state['dados_projeto']['coordenador'])
            st.session_state['dados_projeto']['supervisor'] = st.text_input("Supervisor de Campo", value=st.session_state['dados_projeto']['supervisor'])
        with col_g3:
            st.session_state['dados_projeto']['geologo'] = st.text_input("Geólogo Responsável", value=st.session_state['dados_projeto']['geologo'])
            st.session_state['dados_projeto']['sondador'] = st.text_input("Sondador / Equipe", value=st.session_state['dados_projeto']['sondador'])

    col_furo1, col_furo2 = st.columns(2)
    with col_furo1:
        st.session_state['dados_projeto']['furo_id'] = st.text_input("ID do Furo", value=st.session_state['dados_projeto']['furo_id'])
    with col_furo2:
        diams = ['HQ (63.5mm)', 'NQ (47.6mm)', 'BQ (36.5mm)', 'RC (Circ. Reversa)', 'Outro']
        idx_diam = diams.index(st.session_state['dados_projeto']['diametro']) if st.session_state['dados_projeto']['diametro'] in diams else 0
        st.session_state['dados_projeto']['diametro'] = st.selectbox("Diâmetro", diams, index=idx_diam)

    with st.expander("🌐 Coordenadas GPS e Mapa do Furo", expanded=True):
        st.components.v1.html("""
            <script>
            if (navigator.geolocation) {
                navigator.geolocation.getCurrentPosition(function(position) {
                    const lat = position.coords.latitude;
                    const lon = position.coords.longitude;
                    const urlParams = new URLSearchParams(window.parent.location.search);
                    if (urlParams.get('lat') !== lat.toString() || urlParams.get('lon') !== lon.toString()) {
                        urlParams.set('lat', lat);
                        urlParams.set('lon', lon);
                        window.parent.location.search = urlParams.toString();
                    }
                });
            }
            </script>
        """, height=0)

        params = st.query_params
        if 'lat' in params and 'lon' in params:
            try:
                st.session_state['dados_projeto']['lat_gps'] = float(params['lat'])
                st.session_state['dados_projeto']['lon_gps'] = float(params['lon'])
            except ValueError:
                pass

        col_geo1, col_geo2, col_geo3, col_geo4 = st.columns(4)
        with col_geo1:
            st.session_state['dados_projeto']['lat_gps'] = st.number_input("Latitude", value=st.session_state['dados_projeto']['lat_gps'], format="%.6f")
            st.session_state['dados_projeto']['lon_gps'] = st.number_input("Longitude", value=st.session_state['dados_projeto']['lon_gps'], format="%.6f")
        with col_geo2:
            st.session_state['dados_projeto']['datum'] = st.text_input("Datum", value=st.session_state['dados_projeto']['datum'])
        with col_geo3:
            st.session_state['dados_projeto']['inclinacao'] = st.number_input("Inclinação (°)", value=st.session_state['dados_projeto']['inclinacao'], format="%.1f")
            st.session_state['dados_projeto']['azimute'] = st.number_input("Azimute (°)", value=st.session_state['dados_projeto']['azimute'], format="%.1f")
        with col_geo4:
            st.session_state['dados_projeto']['data_inicio'] = st.date_input("Data de Início", value=st.session_state['dados_projeto']['data_inicio'])
            st.session_state['dados_projeto']['data_fim'] = st.date_input("Data de Término", value=st.session_state['dados_projeto']['data_fim'])

        m = folium.Map(location=[st.session_state['dados_projeto']['lat_gps'], st.session_state['dados_projeto']['lon_gps']], zoom_start=16, tiles=None)
        folium.TileLayer(
            tiles='https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}',
            attr='Esri World Imagery', name='Satélite (Esri)', overlay=False
        ).add_to(m)
        folium.Marker([st.session_state['dados_projeto']['lat_gps'], st.session_state['dados_projeto']['lon_gps']], popup=f"Furo: {st.session_state['dados_projeto']['furo_id']}", icon=folium.Icon(color='red')).add_to(m)
        st_folium(m, width="100%", height=350)

# --- PÁGINA 2: CADASTRO DE MANOBRAS ---
elif menu_opcao == "🛠️ 2. Cadastro de Manobras":
    st.title("🛠️ Registro de Manobras de Campo")

    prox_de = st.session_state['manobras'][-1]['Para (m)'] if st.session_state['manobras'] else 0.0
    prox_para = round(prox_de + 1.5, 2)
    rec_total_ant = st.session_state['manobras'][-1]['Rec. Total (m)'] if st.session_state['manobras'] else 0.0

    st.subheader("🛠️ Peça de Corte e Revestimento")
    col_pc1, col_pc2, col_pc3, col_pc4, col_pc5 = st.columns(5)
    with col_pc1:
        peca_diam = st.text_input("Diâm. Peça", value="NQ")
    with col_pc2:
        peca_coroa = st.text_input("Coroa nº", placeholder="Ex: 89173-17")
    with col_pc3:
        peca_calib = st.text_input("Calib. nº", placeholder="Ex: 1381/17")
    with col_pc4:
        num_caixa = st.number_input("Nº da Caixa", min_value=1, value=1, step=1)
    with col_pc5:
        revest_info = st.text_input("Revestimento", placeholder="Ex: HQ De 0,00 até 34,40m")

    st.markdown("---")
    st.subheader("📊 Avanço e Parâmetros Geotécnicos")
    col_m1, col_m2, col_m3, col_m4, col_m5 = st.columns(5)
    with col_m1:
        de = st.number_input("De (m)", value=float(prox_de), step=0.5, format="%.2f")
    with col_m2:
        para = st.number_input("Para (m)", value=float(prox_para), step=0.5, format="%.2f")
    with col_m3:
        rec = st.number_input("Rec. (m)", value=round(para - de, 2), step=0.1, format="%.2f")
    with col_m4:
        rec_total = st.number_input("Rec. Total Acum. (m)", value=round(rec_total_ant + rec, 2), step=0.1, format="%.2f")
    with col_m5:
        rqd = st.number_input("RQD (m)", value=round((para - de) * 0.8, 2), step=0.1, format="%.2f")

    col_h1, col_h2, col_h3, col_h4 = st.columns(4)
    with col_h1:
        hora_ini = st.time_input("Horário Inicial", value=datetime.now().time())
    with col_h2:
        hora_fim = st.time_input("Horário Final", value=datetime.now().time())
    with col_h3:
        tempo_refeicao = st.text_input("Refeição", placeholder="Ex: 01:00")
    with col_h4:
        manutencao_prev = st.text_input("Manutenção", placeholder="Ex: 00:15")

    col_l1, col_l2 = st.columns(2)
    with col_l1:
        litologia = st.selectbox("Litologia", list(DADOS_LITOLOGIA.keys()))
    with col_l2:
        alteracao = st.selectbox("Alteração", ['Solo / Inconsol.', 'Completamente Alterada', 'Muito Alterada', 'Moderadamente Alterada', 'Pouco Alterada', 'Rocha Sã'])

    st.subheader("📷 Registro Fotográfico da Amostra")
    aba_cam, aba_up = st.tabs(["📸 Tirar Foto Agora", "📁 Carregar da Galeria"])

    img_capturada = None
    with aba_cam:
        foto_cam = st.camera_input("Tirar foto da caixa de testemunho")
        if foto_cam: img_capturada = Image.open(foto_cam)

    with aba_up:
        foto_file = st.file_uploader("Selecione uma imagem", type=['jpg', 'jpeg', 'png'])
        if foto_file and not img_capturada: img_capturada = Image.open(foto_file)

    col_btn1, col_btn2 = st.columns([1, 4])
    with col_btn1:
        btn_adicionar = st.button("➕ Adicionar Manobra", type="primary")
    with col_btn2:
        btn_remover = st.button("🗑️ Remover Última")

    if btn_adicionar:
        avanco = round(para - de, 2)
        if avanco <= 0:
            st.error("⚠️ O valor 'Para' deve ser maior que 'De'!")
        else:
            pct_rec = min(100.0, round((rec / avanco) * 100, 1)) if avanco > 0 else 0.0
            pct_rqd = min(100.0, round((rqd / avanco) * 100, 1)) if avanco > 0 else 0.0
            
            if pct_rqd < 25: rqd_class = 'Muito Pobre'
            elif pct_rqd < 50: rqd_class = 'Pobre'
            elif pct_rqd < 75: rqd_class = 'Razoável'
            elif pct_rqd < 90: rqd_class = 'Boa'
            else: rqd_class = 'Excelente'

            st.session_state['manobras'].append({
                'Manobra': len(st.session_state['manobras']) + 1,
                'De (m)': de, 'Para (m)': para, 'Avanço (m)': avanco,
                'Rec. (m)': rec, 'Rec. Total (m)': rec_total, 'Rec (%)': pct_rec, 
                'RQD (m)': rqd, 'RQD (%)': pct_rqd, 'Qualidade RQD': rqd_class,
                'Diâm. Peça': peca_diam, 'Coroa nº': peca_coroa, 'Calib. nº': peca_calib,
                'Nº Caixa': num_caixa, 'Revestimento': revest_info,
                'Hora Inicial': hora_ini.strftime("%H:%M"), 
                'Hora Final': hora_fim.strftime("%H:%M"),
                'Refeição': tempo_refeicao,
                'Manutenção Preventiva': manutencao_prev,
                'Litologia': litologia, 'Alteração': alteracao, 
                'Foto': img_capturada
            })
            st.success("✅ Manobra registrada!")
            st.rerun()

    if btn_remover and st.session_state['manobras']:
        st.session_state['manobras'].pop()
        st.warning("🗑️ Última manobra removida.")
        st.rerun()

    if st.session_state['manobras']:
        st.markdown("---")
        st.subheader("📋 Tabela de Manobras Cadastradas")
        df_manobras = pd.DataFrame(st.session_state['manobras'])
        st.dataframe(df_manobras.drop(columns=['Foto']), use_container_width=True, hide_index=True)

# --- PÁGINA 3: PERFIL GEOLÓGICO ---
elif menu_opcao == "📊 3. Perfil Geológico":
    st.title("📊 Perfil Litológico e Estratigrafia")

    st.session_state['dados_projeto']['obs_gerais'] = st.text_area(
        "📝 Observações Técnicas Gerais / Notas de Campo do Furo", 
        value=st.session_state['dados_projeto']['obs_gerais'],
        placeholder="Digite observações importantes sobre o furo, trocas de ferramenta, perdas de água, fraturamento especial, etc...",
        height=100
    )

    if st.session_state['manobras']:
        df_manobras = pd.DataFrame(st.session_state['manobras'])

        plt.rcParams['hatch.linewidth'] = 1.2
        plt.rcParams['hatch.color'] = '#333333'
        fig, (ax_lito, ax_rqd, ax_rec) = plt.subplots(1, 3, figsize=(11, 4.5), sharey=True, gridspec_kw={'width_ratios': [1.3, 2, 2]})
        
        prof_max = df_manobras['Para (m)'].max()
        ax_lito.set_ylim(prof_max, 0)

        for _, row in df_manobras.iterrows():
            de_m, para_m, lito = row['De (m)'], row['Para (m)'], row['Litologia']
            rqd_val, rec_val = row['RQD (%)'], row['Rec (%)']
            info_lito = DADOS_LITOLOGIA.get(lito, {'cor': '#808080', 'hatch': ''})

            rect = mpatches.Rectangle((0, de_m), 1, para_m - de_m, facecolor=info_lito['cor'], hatch=info_lito['hatch'], edgecolor='#1E293B', linewidth=1.2)
            ax_lito.add_patch(rect)
            ax_lito.axhline(para_m, color='#0F172A', linestyle='--', linewidth=0.8)
            ax_lito.text(0.5, (de_m + para_m)/2, f"{lito}\n({de_m:.1f}m - {para_m:.1f}m)", ha='center', va='center', fontsize=8, fontweight='bold', bbox=dict(boxstyle='round,pad=0.2', facecolor='#FFFFFF', alpha=0.85, edgecolor='#94A3B8'))

            color_rqd = '#EF4444' if rqd_val < 25 else '#F97316' if rqd_val < 50 else '#EAB308' if rqd_val < 75 else '#3B82F6' if rqd_val < 90 else '#22C55E'
            ax_rqd.barh(y=de_m + (para_m - de_m)/2, width=rqd_val, height=(para_m - de_m)*0.8, color=color_rqd, edgecolor='black', linewidth=0.8)
            ax_rec.barh(y=de_m + (para_m - de_m)/2, width=rec_val, height=(para_m - de_m)*0.8, color='#0284C7', edgecolor='black', linewidth=0.8)

        ax_lito.set_xlim(0, 1)
        ax_lito.set_title("Estratigrafia", fontsize=10, fontweight='bold')
        ax_lito.set_ylabel("Profundidade (m)", fontsize=9, fontweight='bold')
        ax_lito.get_xaxis().set_visible(False)
        ax_rqd.set_xlim(0, 105)
        ax_rqd.set_title("RQD (%)", fontsize=10, fontweight='bold')
        ax_rec.set_xlim(0, 105)
        ax_rec.set_title("Recuperação (%)", fontsize=10, fontweight='bold')
        plt.tight_layout()

        st.pyplot(fig)
        plt.close(fig)
    else:
        st.info("ℹ️ Nenhum dado de manobra cadastrado. Acesse a aba **2. Cadastro de Manobras** para registrar os lançamentos.")

# --- PÁGINA 4: EXPORTAÇÃO & RELATÓRIO PDF ---
elif menu_opcao == "📄 4. Exportação & Relatório PDF":
    st.title("📄 Exportação e Relatório Técnico")

    if not st.session_state['manobras']:
        st.warning("⚠️ Cadastre pelo menos uma manobra para poder gerar os relatórios em Excel e PDF.")
    else:
        df_manobras = pd.DataFrame(st.session_state['manobras'])
        dados = st.session_state['dados_projeto']

        st.markdown("### ✍️ Validação Técnica")
        st.info(f"O documento conterá o campo de assinatura técnica do **{dados['geologo']}**.")

        col_exp1, col_exp2 = st.columns(2)

        # --- EXPORTAÇÃO EXCEL ---
        with col_exp1:
            st.subheader("📊 Planilha Excel")
            buffer_xls = io.BytesIO()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Boletim de Sondagem"
            ws.views.sheetView[0].showGridLines = True

            font_titulo = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
            font_sub = Font(name='Calibri', size=10, italic=True, color='FFFFFF')
            font_sec = Font(name='Calibri', size=11, bold=True, color='0F172A')
            font_header = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
            font_body = Font(name='Calibri', size=10)
            font_total = Font(name='Calibri', size=10, bold=True)

            fill_banner = PatternFill(start_color='0284C7', end_color='0284C7', fill_type='solid')
            fill_sec = PatternFill(start_color='E0F2FE', end_color='E0F2FE', fill_type='solid')
            fill_header = PatternFill(start_color='0369A1', end_color='0369A1', fill_type='solid')
            fill_zebra = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
            fill_total = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')

            thin_border = Border(
                left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
                top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1')
            )
            double_bottom = Border(top=Side(style='thin', color='0F172A'), bottom=Side(style='double', color='0F172A'))

            align_center = Alignment(horizontal='center', vertical='center')
            align_left = Alignment(horizontal='left', vertical='center')
            align_right = Alignment(horizontal='right', vertical='center')

            if dados['logo_img']:
                img_logo_excel_buf = io.BytesIO()
                dados['logo_img'].save(img_logo_excel_buf, format='PNG')
                img_logo_excel_buf.seek(0)
                xl_logo = OpenpyxlImage(img_logo_excel_buf)
                xl_logo.width = 110
                xl_logo.height = 40
                ws.add_image(xl_logo, 'A1')

            ws.merge_cells('C1:L1')
            ws['C1'] = dados['empresa'].upper()
            ws['C1'].font = font_titulo
            ws['C1'].fill = fill_banner
            ws['C1'].alignment = align_center

            ws.merge_cells('A2:L2')
            ws['A2'] = f"BOLETIM TÉCNICO DE SONDAGEM GEOLÓGICA - FURO {dados['furo_id']}"
            ws['A2'].font = font_sub
            ws['A2'].fill = fill_banner
            ws['A2'].alignment = align_center

            ws.merge_cells('A4:L4')
            ws['A4'] = "1. DADOS DE GESTÃO E LOCALIZAÇÃO"
            ws['A4'].font = font_sec
            ws['A4'].fill = fill_sec

            dados_header = [
                [("Projeto:", dados['projeto']), ("Coordenador:", dados['coordenador']), ("Latitude:", dados['lat_gps'])],
                [("ID Furo:", dados['furo_id']), ("Supervisor:", dados['supervisor']), ("Longitude:", dados['lon_gps'])],
                [("Diâmetro:", dados['diametro']), ("Geólogo Resp.:", dados['geologo']), ("Início:", str(dados['data_inicio']))],
                [("Inclin./Az.:", f"{dados['inclinacao']}° / {dados['azimute']}°"), ("Sondador:", dados['sondador']), ("Datum:", dados['datum'])]
            ]

            curr_row = 5
            for row in dados_header:
                col_pairs = [(1,2,3), (4,5,6), (7,8,9)]
                for idx, (lbl, val) in enumerate(row):
                    c_lbl, c_val_start, c_val_end = col_pairs[idx]
                    ws.cell(row=curr_row, column=c_lbl, value=lbl).font = Font(name='Calibri', size=10, bold=True)
                    ws.merge_cells(start_row=curr_row, start_column=c_val_start, end_row=curr_row, end_column=c_val_end)
                    cell_v = ws.cell(row=curr_row, column=c_val_start, value=val)
                    cell_v.font = font_body
                curr_row += 1

            curr_row += 1
            ws.merge_cells(f'A{curr_row}:L{curr_row}')
            ws[f'A{curr_row}'] = "2. REGISTRO DE MANOBRAS E PARÂMETROS GEOTÉCNICOS"
            ws[f'A{curr_row}'].font = font_sec
            ws[f'A{curr_row}'].fill = fill_sec

            curr_row += 1
            df_excel = df_manobras.drop(columns=['Foto'])
            
            for c_idx, col_name in enumerate(df_excel.columns, 1):
                cell = ws.cell(row=curr_row, column=c_idx, value=col_name)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = thin_border

            header_row_idx = curr_row
            curr_row += 1
            for r_idx, row in df_excel.iterrows():
                row_fill = fill_zebra if r_idx % 2 == 1 else PatternFill(fill_type=None)
                for c_idx, val in enumerate(row.values, 1):
                    cell = ws.cell(row=curr_row, column=c_idx, value=val)
                    cell.font = font_body
                    cell.border = thin_border
                    cell.fill = row_fill
                    if isinstance(val, (int, float)):
                        cell.alignment = align_right
                    else:
                        cell.alignment = align_center if c_idx == 1 else align_left
                curr_row += 1

            # Observações Gerais
            curr_row += 2
            ws.merge_cells(f'A{curr_row}:L{curr_row}')
            ws[f'A{curr_row}'] = "3. OBSERVAÇÕES TÉCNICAS E NOTAS DE CAMPO"
            ws[f'A{curr_row}'].font = font_sec
            ws[f'A{curr_row}'].fill = fill_sec

            curr_row += 1
            ws.merge_cells(f'A{curr_row}:L{curr_row+2}')
            ws[f'A{curr_row}'] = dados['obs_gerais'] if dados['obs_gerais'] else "Nenhuma observação complementar."
            ws[f'A{curr_row}'].font = font_body

            # Assinatura
            curr_row += 4
            ws.cell(row=curr_row, column=1, value="_________________________________________").font = font_body
            ws.cell(row=curr_row+1, column=1, value=f"{dados['geologo']} - Geólogo Responsável").font = Font(name='Calibri', size=10, bold=True)

            wb.save(buffer_xls)
            st.download_button(
                label="📊 Baixar Planilha Excel (.xlsx)",
                data=buffer_xls.getvalue(),
                file_name=f"Boletim_{dados['furo_id']}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

        # --- EXPORTAÇÃO PDF RELATÓRIO TÉCNICO ABNT ---
        with col_exp2:
            st.subheader("📄 Relatório PDF Profissional")

            class NumberedCanvas(canvas.Canvas):
                def __init__(self, *args, **kwargs):
                    super().__init__(*args, **kwargs)
                    self._saved_page_states = []

                def showPage(self):
                    self._saved_page_states.append(dict(self.__dict__))
                    self._startPage()

                def save(self):
                    num_pages = len(self._saved_page_states)
                    for state in self._saved_page_states:
                        self.__dict__.update(state)
                        self.draw_page_number(num_pages)
                        super().showPage()
                    super().save()

                def draw_page_number(self, page_count):
                    self.setFont("Times-Roman", 9)
                    text = f"Página {self._pageNumber} de {page_count}"
                    self.drawRightString(19.0 * cm, 1.2 * cm, text)

            pdf_buf = io.BytesIO()
            doc = SimpleDocTemplate(
                pdf_buf, pagesize=portrait(A4),
                leftMargin=3.0*cm, rightMargin=2.0*cm,
                topMargin=3.0*cm, bottomMargin=2.0*cm
            )
            elements = []
            styles = getSampleStyleSheet()

            abnt_titulo_doc = ParagraphStyle('ABNTTituloDoc', parent=styles['Heading1'], fontName='Times-Bold', fontSize=12, leading=14, alignment=1)
            abnt_sec = ParagraphStyle('ABNTSec', parent=styles['Heading2'], fontName='Times-Bold', fontSize=10, leading=12, spaceBefore=8, spaceAfter=4)
            abnt_text = ParagraphStyle('ABNTText', parent=styles['Normal'], fontName='Times-Roman', fontSize=8, leading=10)
            abnt_th = ParagraphStyle('ABNTTH', parent=styles['Normal'], fontName='Times-Bold', fontSize=7.5, leading=9, alignment=1)
            abnt_td = ParagraphStyle('ABNTTD', parent=styles['Normal'], fontName='Times-Roman', fontSize=7.5, leading=9, alignment=1)

            # Cabeçalho do PDF
            if dados['logo_img']:
                img_logo_pdf_buf = io.BytesIO()
                dados['logo_img'].save(img_logo_pdf_buf, format='PNG')
                img_logo_pdf_buf.seek(0)
                rl_logo = RLImage(img_logo_pdf_buf, width=3.5*cm, height=1.2*cm)
                
                header_table = Table([[rl_logo, Paragraph(f"<b>{dados['empresa'].upper()}</b><br/>RELATÓRIO TÉCNICO DE SONDAGEM - FURO <b>{dados['furo_id']}</b>", abnt_titulo_doc)]], colWidths=[4.0*cm, 12.0*cm])
                header_table.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'MIDDLE')]))
                elements.append(header_table)
            else:
                elements.append(Paragraph(f"<b>{dados['empresa'].upper()}</b>", abnt_titulo_doc))
                elements.append(Paragraph(f"RELATÓRIO TÉCNICO DE SONDAGEM — FURO <b>{dados['furo_id']}</b>", abnt_titulo_doc))

            elements.append(Spacer(1, 8))

            # Dados Gerais
            elements.append(Paragraph("<b>1. DADOS DE GESTÃO E LOCALIZAÇÃO</b>", abnt_sec))
            dados_furo_table = [
                [Paragraph("<b>Projeto:</b>", abnt_text), Paragraph(dados['projeto'], abnt_text), Paragraph("<b>Coordenador:</b>", abnt_text), Paragraph(dados['coordenador'], abnt_text)],
                [Paragraph("<b>ID Furo:</b>", abnt_text), Paragraph(dados['furo_id'], abnt_text), Paragraph("<b>Supervisor:</b>", abnt_text), Paragraph(dados['supervisor'], abnt_text)],
                [Paragraph("<b>Diâmetro:</b>", abnt_text), Paragraph(dados['diametro'], abnt_text), Paragraph("<b>Geólogo Resp.:</b>", abnt_text), Paragraph(dados['geologo'], abnt_text)],
                [Paragraph("<b>Período:</b>", abnt_text), Paragraph(f"{dados['data_inicio']} a {dados['data_fim']}", abnt_text), Paragraph("<b>Sondador:</b>", abnt_text), Paragraph(dados['sondador'], abnt_text)],
                [Paragraph("<b>Incl./Az.:</b>", abnt_text), Paragraph(f"{dados['inclinacao']}° / {dados['azimute']}°", abnt_text), Paragraph("<b>Datum/Coord:</b>", abnt_text), Paragraph(f"{dados['datum']} ({dados['lat_gps']:.4f}, {dados['lon_gps']:.4f})", abnt_text)]
            ]
            t_dados = Table(dados_furo_table, colWidths=[2.5*cm, 5.5*cm, 2.5*cm, 5.5*cm])
            t_dados.setStyle(TableStyle([('GRID', (0,0), (-1,-1), 0.5, colors.grey), ('VALIGN', (0,0), (-1,-1), 'MIDDLE')]))
            elements.append(t_dados)

            elements.append(Spacer(1, 8))

            # Inclusão do Mapa no PDF
            elements.append(Paragraph("<b>2. MAPA DE LOCALIZAÇÃO DO FURO (SATÉLITE ESRI)</b>", abnt_sec))
            buf_mapa = obter_mapa_satelite_esri_alta_res(dados['lat_gps'], dados['lon_gps'])
            if buf_mapa:
                rl_mapa = RLImage(buf_mapa, width=16.0*cm, height=6.0*cm)
                elements.append(rl_mapa)
            else:
                elements.append(Paragraph("<i>[Mapa Indisponível]</i>", abnt_text))

            elements.append(Spacer(1, 8))

            # Tabela de Manobras no PDF
            elements.append(Paragraph("<b>3. REGISTRO DE MANOBRAS E PARÂMETROS GEOTÉCNICOS</b>", abnt_sec))
            
            headers_pdf = ["Mnb", "De (m)", "Para (m)", "Rec (m)", "Rec (%)", "RQD (%)", "Litologia", "Alteração"]
            table_pdf_data = [[Paragraph(h, abnt_th) for h in headers_pdf]]

            for m_item in st.session_state['manobras']:
                table_pdf_data.append([
                    Paragraph(str(m_item['Manobra']), abnt_td),
                    Paragraph(f"{m_item['De (m)']:.2f}", abnt_td),
                    Paragraph(f"{m_item['Para (m)']:.2f}", abnt_td),
                    Paragraph(f"{m_item['Rec. (m)']:.2f}", abnt_td),
                    Paragraph(f"{m_item['Rec (%)']:.1f}%", abnt_td),
                    Paragraph(f"{m_item['RQD (%)']:.1f}%", abnt_td),
                    Paragraph(m_item['Litologia'], abnt_td),
                    Paragraph(m_item['Alteração'], abnt_td)
                ])

            t_manobras = Table(table_pdf_data, colWidths=[1.0*cm, 1.8*cm, 1.8*cm, 1.8*cm, 1.8*cm, 1.8*cm, 3.5*cm, 2.5*cm])
            t_manobras.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0369A1')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE')
            ]))
            elements.append(t_manobras)

            elements.append(Spacer(1, 10))

            # Observações e Assinatura Técnica
            elements.append(Paragraph("<b>4. OBSERVAÇÕES TÉCNICAS E ASSINATURA</b>", abnt_sec))
            obs_texto = dados['obs_gerais'] if dados['obs_gerais'] else "Nenhuma observação cadastrada."
            elements.append(Paragraph(f"<b>Notas de Campo:</b> {obs_texto}", abnt_text))

            elements.append(Spacer(1, 15))
            
            assinatura_data = [
                [Paragraph("________________________________________________", abnt_th)],
                [Paragraph(f"<b>{dados['geologo']}</b><br/>Geólogo Responsável", abnt_th)]
            ]
            t_ass = Table(assinatura_data, colWidths=[16.0*cm])
            t_ass.setStyle(TableStyle([('ALIGN', (0,0), (-1,-1), 'CENTER')]))
            elements.append(KeepTogether(t_ass))

            doc.build(elements, canvasmaker=NumberedCanvas)

            st.download_button(
                label="📄 Baixar Relatório PDF (.pdf)",
                data=pdf_buf.getvalue(),
                file_name=f"Relatorio_Sondagem_{dados['furo_id']}.pdf",
                mime="application/pdf",
                use_container_width=True
            )
